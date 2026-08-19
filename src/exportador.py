"""
exportador.py
Escribe DataFrames a .xlsx con formato legible: encabezado resaltado,
ancho de columna automático, encabezado congelado y formato numérico/
fecha según el nombre del campo.

Usa Workbook(write_only=True) en vez de pd.ExcelWriter: con hojas de
decenas de miles de filas, escribir celda por celda vía openpyxl normal
(el motor de to_excel) es el cuello de botella real de todo el programa,
más lento que leer y parsear el archivo plano de origen. El modo
write-only transmite cada fila directamente al XML sin mantener el
worksheet completo en memoria, a costa de no poder releer/reposicionar
celdas ya escritas — por eso el estilo (relleno de encabezado, negrita,
formato de número) se define en el momento de crear cada WriteOnlyCell,
antes de hacer ws.append(fila), en lugar de aplicarse en una pasada
posterior sobre el worksheet ya escrito.
"""

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from utils import logger, manejar_errores

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center")
FORMATO_MONEDA = "#,##0.00"
FORMATO_FECHA = "DD/MM/YYYY"


def _clasificar_columnas(df: pd.DataFrame) -> dict:
    """
    Determina, por nombre de columna, qué formato de número aplica
    (moneda si el nombre contiene "monto"/"valor", fecha si contiene
    "fecha"), igual que la clasificación previa por substring.
    """
    formatos = {}
    for col_name in df.columns:
        nombre = str(col_name).lower()
        if "monto" in nombre or "valor" in nombre:
            formatos[col_name] = FORMATO_MONEDA
        elif "fecha" in nombre:
            formatos[col_name] = FORMATO_FECHA
    return formatos


def _anchos_columnas(df: pd.DataFrame) -> list:
    anchos = []
    for col_name in df.columns:
        # fillna("") antes de astype(str): en pandas 3.x, astype(str) sobre
        # NaN en una columna float (ej. MONTO-1 cuando pd.to_numeric(errors=
        # "coerce") no pudo convertir el valor) NO produce la cadena "nan"
        # como en pandas < 3 — deja el NaN como float, y map(len) revienta
        # con "object of type 'float' has no len()".
        largo_max = df[col_name].fillna("").astype(str).map(len).max() if len(df) else 0
        largo_max = max(largo_max, len(str(col_name)))
        anchos.append(min(largo_max + 3, 40))
    return anchos


def _escribir_hoja_dataframe(wb, nombre_hoja: str, df: pd.DataFrame):
    ws = wb.create_sheet(nombre_hoja)

    formatos = _clasificar_columnas(df)
    anchos = _anchos_columnas(df)

    # En modo write-only, freeze_panes y column_dimensions solo quedan en
    # el .xlsx si se fijan ANTES del primer ws.append(): la hoja se
    # transmite en streaming y esos atributos de cabecera del worksheet
    # ya no se pueden tocar una vez empezaron a salir filas.
    ws.freeze_panes = "A2"
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    fila_header = []
    for col_name in df.columns:
        celda = WriteOnlyCell(ws, value=col_name)
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = HEADER_ALIGN
        fila_header.append(celda)
    ws.append(fila_header)

    columnas = list(df.columns)
    formato_por_columna = [formatos.get(col) for col in columnas]

    # NaN (ej. MONTO-1 no numérico tras pd.to_numeric(errors="coerce"))
    # se pasa como None para que quede como celda vacía, igual que
    # to_excel lo hacía automáticamente.
    valores = df.astype(object).where(df.notna(), None).values.tolist()
    for fila in valores:
        fila_out = []
        for valor, formato in zip(fila, formato_por_columna):
            if formato and valor is not None:
                celda = WriteOnlyCell(ws, value=valor)
                celda.number_format = formato
                fila_out.append(celda)
            else:
                fila_out.append(valor)
        ws.append(fila_out)


def _escribir_hoja_libre(wb, nombre_hoja: str, celdas: list):
    """
    Escribe una hoja de posición libre (no tabular), a partir de una lista
    de celdas: [{"fila": int, "columna": "D", "valor": ..., "negrita": bool,
    "formato_numero": str|None}, ...]. Se usa para replicar layouts fijos
    (ej. la plantilla de VALOR CINTA) en vez de un DataFrame con encabezado.

    write-only no permite direccionar celdas sueltas (ws["D3"]) como antes,
    así que se arma una grilla completa desde la columna A hasta la última
    columna usada (dejando huecos en None) y se manda fila por fila con
    ws.append — las columnas conservan su letra real (D, E, F...) para que
    la plantilla replicada no cambie de posición.
    """
    ws = wb.create_sheet(nombre_hoja)
    if not celdas:
        return

    # Igual que en _escribir_hoja_dataframe: column_dimensions debe fijarse
    # antes del primer ws.append() en modo write-only.
    columnas_usadas = {c["columna"] for c in celdas}
    for col in columnas_usadas:
        ws.column_dimensions[col].width = 24

    max_fila = max(c["fila"] for c in celdas)
    max_col = max(column_index_from_string(c["columna"]) for c in celdas)

    grid = [[None] * max_col for _ in range(max_fila)]
    for celda_def in celdas:
        celda = WriteOnlyCell(ws, value=celda_def["valor"])
        if celda_def.get("negrita"):
            celda.font = Font(bold=True)
        formato = celda_def.get("formato_numero")
        if formato:
            celda.number_format = formato
        col_idx = column_index_from_string(celda_def["columna"]) - 1
        grid[celda_def["fila"] - 1][col_idx] = celda

    for fila in grid:
        ws.append(fila)


@manejar_errores
def guardar_con_formato(path: Path, hojas: dict) -> bool:
    """
    hojas: {"NombreHoja": dataframe_o_lista_de_celdas, ...}
    Cada hoja puede ser un DataFrame (tabla con encabezado, formato
    automático de moneda/fecha) o una lista de celdas de posición libre
    (ver _escribir_hoja_libre), para layouts fijos que no son tabulares.
    Si algún DataFrame viene vacío (0 filas), igual se escribe la hoja
    con encabezados en lugar de fallar, pero se avisa explícitamente
    en consola y en el log — así el usuario no pasa por alto que un
    filtro no encontró resultados.

    Devuelve True si el archivo se generó correctamente. Si algo falla
    (ej. el archivo está abierto en Excel), @manejar_errores captura el
    error y esta función devuelve None — el llamador debe revisar el
    resultado antes de asumir éxito, para no mostrar un mensaje de
    "generado" después de un fallo real.
    """
    hojas_vacias = [
        nombre for nombre, contenido in hojas.items()
        if isinstance(contenido, pd.DataFrame) and len(contenido) == 0
    ]

    wb = openpyxl.Workbook(write_only=True)
    for nombre_hoja, contenido in hojas.items():
        if isinstance(contenido, pd.DataFrame):
            _escribir_hoja_dataframe(wb, nombre_hoja, contenido)
        else:
            _escribir_hoja_libre(wb, nombre_hoja, contenido)
    wb.save(path)

    print(f"[OK] Generado: {path}")
    for nombre_hoja in hojas_vacias:
        mensaje = f"   [!]  '{nombre_hoja}' en {path.name} quedó sin filas — revisa el filtro aplicado."
        print(mensaje)
        logger.warning(f"Hoja vacía: {nombre_hoja} en {path}")

    return True
